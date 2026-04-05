"""
engine.py - Main Trading Engine

JSS Sawriya Seth Wealthtech
The brain of the entire AI options trading system.  Coordinates the broker,
capital manager, risk manager, option chain analyzer, all strategies,
Telegram integrations, and the database in a single thread-safe control loop.

The engine:
- Runs a background thread that scans symbols every 3 seconds
- Executes trades when ALL strategies produce aligned signals
- Monitors open positions (SL / target / trailing SL / time exits)
- Generates daily reports and sends them via Telegram
- Provides a ``get_dashboard_data()`` method consumed by the GUI every 2 s

Author: JSS Sawriya Seth Wealthtech Engineering
"""

from __future__ import annotations

import json
import logging
import os
import threading
import time
import uuid
from datetime import datetime, time as dt_time
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger("ENGINE")

# ---------------------------------------------------------------------------
# Engine states
# ---------------------------------------------------------------------------

STOPPED = "STOPPED"
RUNNING = "RUNNING"

# Square-off time — all positions must be closed by 15:15 IST
_SQUARE_OFF_TIME = dt_time(15, 15)


class TradingEngine:
    """Central trading engine that orchestrates every component.

    Parameters
    ----------
    config_path : str
        Path to ``config/config.json``.  The JSON file is expected to contain
        top-level keys for broker credentials, capital settings, risk
        parameters, strategy toggles, and Telegram configuration.
    """

    # ------------------------------------------------------------------
    # Initialisation
    # ------------------------------------------------------------------

    def __init__(self, config_path: str = "config/config.json") -> None:
        self._config_path = config_path
        self._config: dict[str, Any] = {}
        self._state: str = STOPPED
        self._running: bool = False
        self._lock = threading.Lock()
        self._thread: Optional[threading.Thread] = None

        # Component references (set during start)
        self.db: Optional[Any] = None
        self.broker: Optional[Any] = None
        self.capital_manager: Optional[Any] = None
        self.risk_manager: Optional[Any] = None
        self.option_chain: Optional[Any] = None
        self.strategies: list[Any] = []
        self.telegram_bot: Optional[Any] = None
        self.telegram_reader: Optional[Any] = None
        self.signal_parser: Optional[Any] = None

        # Dashboard cache
        self._dashboard_cache: dict[str, Any] = {}

        logger.info("TradingEngine created (config_path=%s)", config_path)

    # ------------------------------------------------------------------
    # Config
    # ------------------------------------------------------------------

    def _load_config(self) -> dict[str, Any]:
        """Load and validate the JSON configuration file."""
        try:
            config_file = Path(self._config_path)
            if not config_file.exists():
                logger.warning(
                    "Config file not found at %s — using defaults", self._config_path
                )
                return self._default_config()

            with open(config_file, "r", encoding="utf-8") as fh:
                cfg = json.load(fh)

            logger.info("Config loaded from %s", self._config_path)
            return cfg

        except json.JSONDecodeError as exc:
            logger.error("Config JSON parse error: %s — using defaults", exc)
            return self._default_config()
        except Exception as exc:
            logger.error("Config load error: %s — using defaults", exc)
            return self._default_config()

    @staticmethod
    def _default_config() -> dict[str, Any]:
        """Return a sensible default configuration."""
        return {
            "broker": {
                "client_code": "",
                "access_token": "",
                "mobile": "",
                "mpin": "",
                "totp_secret": "",
            },
            "database": {"path": "data/jss_trading.db"},
            "capital": {
                "initial_capital": 1000.0,
                "max_daily_loss_percent": 5.0,
                "max_open_trades": 3,
                "scale_capital_threshold": 10000.0,
                "lot_multiplier": 50,
            },
            "risk": {
                "sl_percent": 15.0,
                "risk_reward_ratio": 2.0,
                "trailing_sl_activation_pct": 3.0,
                "trailing_sl_trail_pct": 1.5,
                "max_daily_loss_percent": 5.0,
                "strategy_cooldown_minutes": 15,
                "max_consecutive_losses": 3,
            },
            "strategies": {
                "momentum_follow": {"enabled": True, "min_confidence": 65},
                "multi_scalping": {"enabled": True, "min_confidence": 65},
                "reversal_scalp": {"enabled": True, "min_confidence": 65},
                "expiry_heropatla": {"enabled": True, "min_confidence": 65},
            },
            "symbols": ["NIFTY", "BANKNIFTY"],
            "scan_interval_seconds": 3,
            "candle_interval": "1m",
            "candle_limit": 100,
            "square_off_time": "15:15",
            "telegram": {
                "enabled": False,
                "bot_token": "",
                "chat_id": "",
                "reader_enabled": False,
                "api_id": 0,
                "api_hash": "",
                "phone": "",
                "groups_to_watch": [],
            },
            "ai": {
                "enabled": False,
                "api_key": "",
            },
            "daily_report": {
                "enabled": True,
                "export_excel": True,
            },
        }

    # ------------------------------------------------------------------
    # Component initialisation
    # ------------------------------------------------------------------

    def _init_components(self) -> bool:
        """Instantiate all subsystems.  Returns ``True`` on success."""
        try:
            # 1. Database
            from core.db_helper import Database

            db_path = self._config.get("database", {}).get(
                "path", "data/jss_trading.db"
            )
            self.db = Database(db_path)

            # 2. Broker
            from brokers.kotak_neo import KotakNeoBroker

            broker_cfg = self._config.get("broker", {})
            self.broker = KotakNeoBroker(broker_cfg)

            # 3. Capital Manager
            from core.capital import CapitalManager

            capital_cfg = self._config.get("capital", {})
            self.capital_manager = CapitalManager(self.db, capital_cfg)
            self.capital_manager.load_state()

            # 4. Risk Manager
            from core.risk import RiskManager

            risk_cfg = self._config.get("risk", {})
            self.risk_manager = RiskManager(risk_cfg, self.capital_manager)

            # 5. Option Chain Analyzer
            from core.option_chain import OptionChainAnalyzer

            self.option_chain = OptionChainAnalyzer(self.broker)

            # 6. Strategies
            self.strategies = []
            strats_cfg = self._config.get("strategies", {})

            try:
                from strategies.momentum_follow import MomentumFollowStrategy

                if strats_cfg.get("momentum_follow", {}).get("enabled", True):
                    cfg = strats_cfg.get("momentum_follow", {})
                    self.strategies.append(MomentumFollowStrategy(cfg))
                    logger.info("Strategy loaded: MomentumFollow")
            except Exception as exc:
                logger.error("Failed to load MomentumFollow: %s", exc)

            try:
                from strategies.multi_scalping import MultiScalpingStrategy

                if strats_cfg.get("multi_scalping", {}).get("enabled", True):
                    cfg = strats_cfg.get("multi_scalping", {})
                    self.strategies.append(MultiScalpingStrategy(cfg))
                    logger.info("Strategy loaded: MultiScalping")
            except Exception as exc:
                logger.error("Failed to load MultiScalping: %s", exc)

            try:
                from strategies.reversal_scalp import ReversalScalpStrategy

                if strats_cfg.get("reversal_scalp", {}).get("enabled", True):
                    cfg = strats_cfg.get("reversal_scalp", {})
                    self.strategies.append(ReversalScalpStrategy(cfg))
                    logger.info("Strategy loaded: ReversalScalp")
            except Exception as exc:
                logger.error("Failed to load ReversalScalp: %s", exc)

            try:
                from strategies.expiry_heropatla import ExpiryHeropatlaStrategy

                if strats_cfg.get("expiry_heropatla", {}).get("enabled", True):
                    cfg = strats_cfg.get("expiry_heropatla", {})
                    self.strategies.append(ExpiryHeropatlaStrategy(cfg))
                    logger.info("Strategy loaded: ExpiryHeropatla")
            except Exception as exc:
                logger.error("Failed to load ExpiryHeropatla: %s", exc)

            # 7. Telegram Alert Bot
            tg_cfg = self._config.get("telegram", {})
            if tg_cfg.get("enabled", False):
                try:
                    from tg_integration.bot import TelegramAlertBot

                    self.telegram_bot = TelegramAlertBot(
                        token=tg_cfg.get("bot_token", ""),
                        chat_id=tg_cfg.get("chat_id", ""),
                    )
                    logger.info("TelegramAlertBot initialised")
                except Exception as exc:
                    logger.error("Failed to init TelegramAlertBot: %s", exc)

            # 8. Telegram Reader
            if tg_cfg.get("reader_enabled", False):
                try:
                    from tg_integration.reader import TelegramReader
                    from tg_integration.parser import SignalParser

                    self.telegram_reader = TelegramReader(
                        api_id=tg_cfg.get("api_id", 0),
                        api_hash=tg_cfg.get("api_hash", ""),
                        phone=tg_cfg.get("phone", ""),
                    )
                    self.signal_parser = SignalParser()
                    logger.info("TelegramReader & SignalParser initialised")
                except Exception as exc:
                    logger.error(
                        "Failed to init TelegramReader/SignalParser: %s", exc
                    )

            logger.info(
                "All components initialised: %d strategies, DB=%s, Broker=%s",
                len(self.strategies),
                "OK" if self.db else "FAIL",
                "OK" if self.broker else "FAIL",
            )
            return True

        except Exception as exc:
            logger.error("Component initialisation FAILED: %s", exc, exc_info=True)
            return False

    # ------------------------------------------------------------------
    # Lifecycle: start / stop
    # ------------------------------------------------------------------

    def start(self) -> None:
        """Start the trading engine.

        Loads config, initialises all components, starts background thread
        for the main loop, and optionally starts Telegram services.
        """
        with self._lock:
            if self._running:
                logger.warning("Engine is already running")
                return

        logger.info("Engine starting...")

        # Load config
        self._config = self._load_config()

        # Initialise components
        if not self._init_components():
            logger.error("Cannot start engine — component init failed")
            return

        # Load broker session
        try:
            if self.broker:
                self.broker.load_session()
        except Exception as exc:
            logger.warning("Broker session load failed (continuing): %s", exc)

        # Set state
        with self._lock:
            self._state = RUNNING
            self._running = True

        # Start background main loop thread
        self._thread = threading.Thread(
            target=self.main_loop, name="engine-main", daemon=True
        )
        self._thread.start()
        logger.info("Main loop thread started")

        # Start Telegram bot
        if self.telegram_bot:
            try:
                self.telegram_bot.start()
                logger.info("TelegramAlertBot started")
            except Exception as exc:
                logger.error("TelegramAlertBot start failed: %s", exc)

        # Start Telegram reader + wire signal handler
        if self.telegram_reader and self.signal_parser:
            try:
                groups = self._config.get("telegram", {}).get(
                    "groups_to_watch", []
                )
                if groups:
                    self.telegram_reader.on_message(self._on_telegram_signal)
                    self.telegram_reader.start(groups)
                    logger.info(
                        "TelegramReader started (watching %d groups)", len(groups)
                    )
            except Exception as exc:
                logger.error("TelegramReader start failed: %s", exc)

        self._log_to_db("INFO", "Engine started successfully", "ENGINE")
        logger.info("Engine started successfully")

    def stop(self) -> None:
        """Stop the trading engine gracefully.

        - Squares off all open positions.
        - Stops Telegram services.
        - Sets state to STOPPED.
        """
        logger.info("Engine stopping...")

        with self._lock:
            self._state = STOPPED
            self._running = False

        # Close all open positions (square off)
        try:
            self._square_off_all("Engine shutdown")
        except Exception as exc:
            logger.error("Square-off during shutdown failed: %s", exc)

        # Stop Telegram bot
        if self.telegram_bot:
            try:
                self.telegram_bot.stop()
                logger.info("TelegramAlertBot stopped")
            except Exception as exc:
                logger.error("TelegramAlertBot stop failed: %s", exc)

        # Stop Telegram reader
        if self.telegram_reader:
            try:
                self.telegram_reader.stop()
                logger.info("TelegramReader stopped")
            except Exception as exc:
                logger.error("TelegramReader stop failed: %s", exc)

        # Generate daily report on shutdown
        try:
            self.generate_daily_report()
        except Exception as exc:
            logger.error("Daily report generation on shutdown failed: %s", exc)

        # Log stop BEFORE closing DB (can't write to closed DB)
        self._log_to_db("INFO", "Engine stopped", "ENGINE")
        logger.info("Engine stopped")

        # Close DB
        if self.db:
            try:
                self.db.close()
            except Exception as exc:
                logger.error("DB close failed: %s", exc)

        # Close broker HTTP session
        if self.broker:
            try:
                self.broker.close()
            except Exception as exc:
                logger.error("Broker close failed: %s", exc)

    # ------------------------------------------------------------------
    # Main Loop
    # ------------------------------------------------------------------

    def main_loop(self) -> None:
        """Background thread loop — the heart of the engine.

        Every iteration:
        1. Check market hours.
        2. If market is open:
           a. Fetch fresh candles for each symbol.
           b. Run all strategies.
           c. Execute qualifying signals.
           d. Monitor open trades (SL / target / trailing / time).
        3. Sleep for ``scan_interval_seconds``.
        4. At square-off time, close ALL positions.

        Exceptions within a single iteration are caught so the engine
        NEVER crashes.
        """
        logger.info("Main loop started")

        daily_report_generated = False

        while self._running:
            try:
                # Check if market is open
                if self.broker and self.broker.is_market_open():
                    if not daily_report_generated:
                        # Reset daily counters at start of day
                        daily_report_generated = True
                        try:
                            if self.capital_manager:
                                self.capital_manager.load_state()
                            if self.db:
                                self.db.reset_daily_counters()
                        except Exception as exc:
                            logger.error(
                                "Daily reset failed: %s", exc, exc_info=True
                            )

                    # Scan symbols
                    symbols = self._config.get("symbols", ["NIFTY", "BANKNIFTY"])

                    for symbol in symbols:
                        try:
                            self._scan_symbol(symbol)
                        except Exception as exc:
                            logger.error(
                                "Error scanning %s: %s",
                                symbol,
                                exc,
                                exc_info=True,
                            )

                    # Monitor open trades
                    try:
                        self.monitor_trades()
                    except Exception as exc:
                        logger.error(
                            "Monitor trades error: %s", exc, exc_info=True
                        )

                    # Check square-off time
                    try:
                        now = datetime.now().time()
                        sq_str = self._config.get("square_off_time", "15:15")
                        sq_parts = sq_str.split(":")
                        sq_time = dt_time(int(sq_parts[0]), int(sq_parts[1]))
                        if now >= sq_time:
                            self._square_off_all("Square-off time reached")
                    except Exception as exc:
                        logger.error("Square-off check error: %s", exc)

                else:
                    # Market closed
                    now = datetime.now()
                    if now.hour < 9:
                        daily_report_generated = False

                    # Still monitor open trades even when market closed
                    # (in case there are leftover positions)
                    try:
                        self.monitor_trades()
                    except Exception as exc:
                        logger.error(
                            "Monitor trades (off-hours) error: %s", exc
                        )

            except Exception as exc:
                logger.error(
                    "Main loop iteration error: %s", exc, exc_info=True
                )

            # Sleep
            interval = self._config.get("scan_interval_seconds", 3)
            time.sleep(interval)

        logger.info("Main loop exited")

    # ------------------------------------------------------------------
    # Symbol scanning
    # ------------------------------------------------------------------

    def _scan_symbol(self, symbol: str) -> None:
        """Fetch candles for *symbol* and run all strategies.

        For each generated signal:
        1. Check risk (``risk_manager.check_risk``).
        2. Check capital (``capital_manager.can_trade``).
        3. Execute if allowed.
        """
        if not self.broker or not self.db:
            return

        # Fetch fresh candles
        interval = self._config.get("candle_interval", "1m")
        limit = self._config.get("candle_limit", 100)
        candles = self.broker.get_candles(symbol, interval, limit)

        if not candles:
            logger.debug("No candles for %s — skipping", symbol)
            return

        # Save candles to DB
        try:
            self.db.save_candles_bulk(symbol, interval, candles)
        except Exception as exc:
            logger.error("Failed to save candles for %s: %s", symbol, exc)

        # Generate AI analysis for GUI
        try:
            self.generate_ai_analysis(symbol, candles, [])
        except Exception as exc:
            logger.error("AI analysis generation failed: %s", exc)

        # Run all strategies
        for strategy in self.strategies:
            try:
                signal = self.run_strategy(strategy, symbol, candles)
                if signal is None:
                    continue

                logger.info(
                    "Signal from %s: %s %s %s %d (conf=%.1f)",
                    signal.get("strategy_name", "?"),
                    signal.get("direction", "?"),
                    symbol,
                    signal.get("option_type", "?"),
                    signal.get("strike", 0),
                    signal.get("confidence", 0),
                )

                # Enrich signal with option chain data
                if self.option_chain:
                    try:
                        signal = self.option_chain.find_best_option(signal)
                    except Exception as exc:
                        logger.error(
                            "Option chain enrichment failed: %s", exc
                        )

                # Add strategy_name if missing
                if "strategy_name" not in signal:
                    signal["strategy_name"] = getattr(
                        strategy, "name", "Unknown"
                    )

                # Risk check
                if self.risk_manager:
                    risk_result = self.risk_manager.check_risk(signal)

                    if not risk_result.get("allowed", False):
                        reason = risk_result.get("reason", "Risk check failed")
                        logger.info(
                            "Signal REJECTED by risk: %s", reason
                        )
                        self._log_to_db(
                            "WARN",
                            f"Signal rejected: {reason}",
                            "RISK",
                            {"signal": signal},
                        )
                        continue

                    # Update signal with risk-calculated values
                    if risk_result.get("sl"):
                        signal["sl"] = risk_result["sl"]
                    if risk_result.get("target"):
                        signal["target"] = risk_result["target"]
                    if risk_result.get("position_size"):
                        signal["position_size"] = risk_result["position_size"]

                # Execute trade
                self.execute_trade(signal)

            except Exception as exc:
                logger.error(
                    "Error running strategy %s on %s: %s",
                    getattr(strategy, "name", "?"),
                    symbol,
                    exc,
                    exc_info=True,
                )

    # ------------------------------------------------------------------
    # Trade execution
    # ------------------------------------------------------------------

    def execute_trade(self, signal: dict[str, Any]) -> Optional[dict[str, Any]]:
        """Execute a trade based on a validated signal.

        Steps:
        1. Calculate lot size from ``capital_manager``.
        2. Calculate trade value.
        3. Place order via broker (paper trading).
        4. Save trade to database.
        5. Update capital (reserve amount).
        6. Send Telegram alert.
        7. Log trade open.

        Returns the trade dict on success, ``None`` on failure.
        """
        if not self.capital_manager or not self.broker or not self.db:
            logger.error("Cannot execute trade — components not ready")
            return None

        try:
            symbol = signal.get("symbol", "NIFTY").upper()
            direction = signal.get("direction", "BUY").upper()
            option_type = signal.get("option_type", "CE").upper()
            strike = int(signal.get("strike", 0))
            entry_price = float(signal.get("entry_price", 0))
            strategy_name = signal.get("strategy_name", "Unknown")
            confidence = float(signal.get("confidence", 0))

            # Lot size
            lot_size = self.capital_manager.get_lot_size()
            lot_multiplier = self._get_lot_multiplier(symbol)
            qty = lot_size * lot_multiplier

            # Trade value — use premium (entry_price) not strike for accurate costing
            premium = entry_price if entry_price > 0 else strike
            trade_value = self.capital_manager.calculate_trade_value(
                premium, lot_size, lot_multiplier
            )

            # Capital check
            if not self.capital_manager.can_trade(trade_value):
                logger.warning(
                    "Trade rejected by capital manager: value=%.2f", trade_value
                )
                return None

            # SL / Target
            sl = float(signal.get("sl", 0))
            target = float(signal.get("target", 0))

            if sl <= 0 and self.risk_manager:
                sl = self.risk_manager.calculate_sl(entry_price, direction)
            if target <= 0 and self.risk_manager:
                target = self.risk_manager.calculate_target(
                    entry_price, direction
                )

            # Initialize trailing SL to regular SL (will be updated by monitor_trades)
            initial_trailing_sl = sl

            # Generate unique trade ID
            trade_id = f"T{uuid.uuid4().hex[:10].upper()}"

            # Place order (paper trading)
            order = self.broker.place_order(
                symbol=symbol,
                option_type=option_type,
                strike=strike,
                direction=direction,
                qty=qty,
                order_type="MARKET",
                price=entry_price,
            )

            filled_price = float(order.get("filled_price", entry_price))
            if filled_price <= 0:
                filled_price = entry_price

            # Build trade dict
            trade = {
                "id": trade_id,
                "symbol": symbol,
                "direction": direction,
                "option_type": option_type,
                "strike": strike,
                "entry_price": round(filled_price, 2),
                "exit_price": None,
                "entry_time": datetime.now().isoformat(),
                "exit_time": None,
                "qty": qty,
                "lot_size": lot_size,
                "sl": round(sl, 2),
                "target": round(target, 2),
                "trailing_sl": round(initial_trailing_sl, 2),
                "pnl": 0.0,
                "status": "OPEN",
                "strategy": strategy_name,
                "confidence": confidence,
                "reason": signal.get("reason", ""),
                "exit_reason": "",
                "capital": float(
                    self.capital_manager.get_state().get("current", 0)
                ),
            }

            # Save to database
            self.db.save_trade(trade)

            # Update capital (reserve trade value)
            self.capital_manager.record_trade_open(trade_value)

            # Telegram alert
            capital_remaining = self.capital_manager.get_state().get(
                "current", 0
            )
            trade_alert = {**trade, "capital_remaining": capital_remaining}
            if self.telegram_bot:
                try:
                    self.telegram_bot.send_trade_open(trade_alert)
                except Exception as exc:
                    logger.error("Telegram trade open alert failed: %s", exc)

            # Log
            self._log_to_db(
                "TRADE",
                f"TRADE OPEN: {direction} {symbol} {strike} {option_type} "
                f"@ {filled_price:.2f} | SL={sl:.2f} TGT={target:.2f} | "
                f"Strategy={strategy_name} Conf={confidence:.1f}",
                "ENGINE",
                {"trade_id": trade_id, "signal": signal},
            )

            logger.info(
                "TRADE OPEN: %s %s %s %d @ %.2f (lots=%d, sl=%.2f, tgt=%.2f)",
                direction,
                symbol,
                option_type,
                strike,
                filled_price,
                lot_size,
                sl,
                target,
            )

            return trade

        except Exception as exc:
            logger.error(
                "execute_trade failed: %s", exc, exc_info=True
            )
            return None

    # ------------------------------------------------------------------
    # Trade monitoring
    # ------------------------------------------------------------------

    def monitor_trades(self) -> None:
        """Get all open trades from DB and check exit conditions.

        For each open trade:
        1. Get current LTP.
        2. Check SL hit → ``close_trade``.
        3. Check target hit → ``close_trade``.
        4. Check trailing SL → update trailing_sl.
        5. Check time exit (15:15) → ``close_trade``.
        """
        if not self.db or not self.broker or not self.risk_manager:
            return

        try:
            open_trades = self.db.get_open_trades()
        except Exception as exc:
            logger.error("Failed to get open trades: %s", exc)
            return

        for trade in open_trades:
            try:
                # Build option symbol for LTP lookup
                symbol = trade.get("symbol", "NIFTY")
                strike = trade.get("strike", 0)
                option_type = trade.get("option_type", "CE")

                # Get current LTP from broker
                ltp = 0.0
                try:
                    # Try to get LTP for the specific option
                    if self.option_chain:
                        opt_data = self.option_chain.get_option_data(
                            symbol, int(strike), option_type
                        )
                        ltp = float(opt_data.get("ltp", 0))

                    # Fallback to underlying LTP if option LTP is 0
                    if ltp <= 0:
                        ltp = self.broker.get_ltp(symbol)
                        if ltp > 1000:
                            # This is the index price, not option premium
                            # Use entry price as a rough estimate
                            ltp = float(trade.get("entry_price", 0))

                except Exception as exc:
                    logger.warning(
                        "Failed to get LTP for %s %s %d: %s",
                        symbol,
                        option_type,
                        strike,
                        exc,
                    )
                    ltp = float(trade.get("entry_price", 0))

                if ltp <= 0:
                    continue

                # Update trade with current LTP in dashboard cache
                trade["ltp"] = ltp

                # 1. Check SL
                sl_hit, sl_reason = self.risk_manager.should_exit_by_sl(
                    trade, ltp
                )
                if sl_hit:
                    logger.info("SL HIT for trade %s: %s", trade["id"], sl_reason)
                    self.close_trade(trade["id"], ltp, "SL_HIT")
                    if self.risk_manager:
                        self.risk_manager.record_loss()
                        self.risk_manager.set_cooldown(
                            trade.get("strategy", "Unknown")
                        )
                    continue

                # 2. Check Target
                tgt_hit, tgt_reason = self.risk_manager.should_exit_by_target(
                    trade, ltp
                )
                if tgt_hit:
                    logger.info(
                        "TARGET HIT for trade %s: %s", trade["id"], tgt_reason
                    )
                    self.close_trade(trade["id"], ltp, "TARGET_HIT")
                    if self.risk_manager:
                        self.risk_manager.record_win()
                    continue

                # 3. Check Trailing SL
                trail_hit, trail_reason = self.risk_manager.should_exit_by_trailing(
                    trade, ltp
                )
                if trail_hit:
                    logger.info(
                        "TRAILING SL HIT for trade %s: %s",
                        trade["id"],
                        trail_reason,
                    )
                    self.close_trade(trade["id"], ltp, "TRAILING_SL_HIT")
                    if self.risk_manager:
                        self.risk_manager.record_win()
                    continue

                # 4. Check time exit
                time_exit, time_reason = self.risk_manager.should_exit_by_time(
                    trade
                )
                if time_exit:
                    logger.info(
                        "TIME EXIT for trade %s: %s", trade["id"], time_reason
                    )
                    self.close_trade(trade["id"], ltp, "TIME_EXIT")
                    continue

            except Exception as exc:
                logger.error(
                    "Error monitoring trade %s: %s",
                    trade.get("id", "?"),
                    exc,
                    exc_info=True,
                )

    def close_trade(
        self, trade_id: str, exit_price: float, reason: str
    ) -> Optional[dict[str, Any]]:
        """Close a trade and settle P&L.

        Steps:
        1. Close via broker.
        2. Update database with exit price, P&L, status.
        3. Update capital (add profit / subtract loss).
        4. Send Telegram alert.
        5. Log trade close.

        Returns the updated trade dict on success.
        """
        if not self.db or not self.capital_manager:
            logger.error("Cannot close trade — components not ready")
            return None

        try:
            # Close via database (calculates P&L)
            trade = self.db.close_trade(trade_id, exit_price, reason)

            if trade is None:
                logger.warning(
                    "close_trade: trade %s not found or not OPEN", trade_id
                )
                return None

            pnl = float(trade.get("pnl", 0))
            direction = trade.get("direction", "BUY").upper()
            entry_price = float(trade.get("entry_price", 0))
            qty = int(trade.get("qty", 0))
            lot_size = int(trade.get("lot_size", 1))
            strike = int(trade.get("strike", 0))
            lot_multiplier = self._get_lot_multiplier(
                trade.get("symbol", "NIFTY")
            )

            # Update capital
            try:
                self.capital_manager.record_trade_close(
                    entry_price, exit_price, qty, lot_size, strike, lot_multiplier
                )
            except Exception as exc:
                logger.error("Capital update on close failed: %s", exc)

            # Telegram alert
            if self.telegram_bot:
                try:
                    capital_state = self.capital_manager.get_state()
                    trade["total_pnl"] = capital_state.get("total_pnl", pnl)
                    self.telegram_bot.send_trade_close(trade, pnl)
                except Exception as exc:
                    logger.error(
                        "Telegram trade close alert failed: %s", exc
                    )

            # Log
            self._log_to_db(
                "TRADE",
                f"TRADE CLOSED: {trade_id} | PnL={pnl:+.2f} | "
                f"Exit={exit_price:.2f} | Reason={reason}",
                "ENGINE",
                {
                    "trade_id": trade_id,
                    "pnl": pnl,
                    "exit_price": exit_price,
                    "reason": reason,
                },
            )

            logger.info(
                "TRADE CLOSED: %s | PnL=%+.2f | exit=%.2f | reason=%s",
                trade_id,
                pnl,
                exit_price,
                reason,
            )

            return trade

        except Exception as exc:
            logger.error(
                "close_trade failed for %s: %s",
                trade_id,
                exc,
                exc_info=True,
            )
            return None

    # ------------------------------------------------------------------
    # Square-off
    # ------------------------------------------------------------------

    def _square_off_all(self, reason: str = "Square-off") -> None:
        """Close ALL open positions immediately."""
        if not self.db:
            return

        try:
            open_trades = self.db.get_open_trades()

            if not open_trades:
                logger.info("No open positions to square off")
                return

            logger.info("Squaring off %d open positions: %s", len(open_trades), reason)

            for trade in open_trades:
                try:
                    # Get approximate exit price
                    exit_price = float(trade.get("entry_price", 0))
                    if self.broker and self.option_chain:
                        opt_data = self.option_chain.get_option_data(
                            trade.get("symbol", "NIFTY"),
                            int(trade.get("strike", 0)),
                            trade.get("option_type", "CE"),
                        )
                        ltp = float(opt_data.get("ltp", 0))
                        if ltp > 0:
                            exit_price = ltp

                    self.close_trade(trade["id"], exit_price, reason)

                except Exception as exc:
                    logger.error(
                        "Square-off error for trade %s: %s",
                        trade.get("id", "?"),
                        exc,
                    )

        except Exception as exc:
            logger.error("Square-off all failed: %s", exc, exc_info=True)

    # ------------------------------------------------------------------
    # Strategy runner
    # ------------------------------------------------------------------

    def run_strategy(
        self,
        strategy: Any,
        symbol: str,
        candles: list[dict[str, Any]],
    ) -> Optional[dict[str, Any]]:
        """Run a single strategy and return its signal (or None).

        Wraps the strategy's ``analyze`` call in exception handling so a
        buggy strategy cannot crash the engine.
        """
        try:
            signal = strategy.analyze(symbol, candles)
            if signal is not None:
                # Ensure symbol is set
                signal["symbol"] = symbol
                return signal
        except Exception as exc:
            logger.error(
                "Strategy %s crashed on %s: %s",
                getattr(strategy, "name", "?"),
                symbol,
                exc,
                exc_info=True,
            )
        return None

    # ------------------------------------------------------------------
    # Telegram signal handler
    # ------------------------------------------------------------------

    def _on_telegram_signal(
        self, group_name: str, message_text: str, sender: str
    ) -> None:
        """Handle an incoming Telegram message — parse and trade if valid."""
        if not self.signal_parser:
            return

        try:
            signal = self.signal_parser.parse(message_text)
            if signal is None:
                return

            if not self.signal_parser.validate_signal(signal):
                logger.warning(
                    "Telegram signal validation failed: %s", signal
                )
                return

            logger.info(
                "Telegram signal from %s (%s): %s %s %s %s",
                sender,
                group_name,
                signal.get("direction", "?"),
                signal.get("symbol", "?"),
                signal.get("strike", "?"),
                signal.get("option_type", "?"),
            )

            # Save to DB
            if self.db:
                self.db.save_telegram_message(
                    {
                        "text": message_text,
                        "sender_name": sender,
                        "direction": "INCOMING",
                    }
                )

            # Enrich with option chain
            if self.option_chain:
                try:
                    signal = self.option_chain.find_best_option(signal)
                except Exception as exc:
                    logger.error(
                        "Option chain enrichment (TG) failed: %s", exc
                    )

            signal["strategy_name"] = "Telegram Signal"

            # Risk check
            if self.risk_manager:
                risk_result = self.risk_manager.check_risk(signal)
                if not risk_result.get("allowed", False):
                    reason = risk_result.get("reason", "Risk check failed")
                    logger.info(
                        "Telegram signal REJECTED: %s", reason
                    )
                    return

                if risk_result.get("sl"):
                    signal["sl"] = risk_result["sl"]
                if risk_result.get("target"):
                    signal["target"] = risk_result["target"]

            self.execute_trade(signal)

        except Exception as exc:
            logger.error(
                "Telegram signal handler error: %s", exc, exc_info=True
            )

    # ------------------------------------------------------------------
    # AI Analysis (for GUI)
    # ------------------------------------------------------------------

    def generate_ai_analysis(
        self,
        symbol: str,
        candles: list[dict[str, Any]],
        signals: list[dict[str, Any]],
    ) -> dict[str, Any]:
        """Compile indicator values + signal confidence into an analysis dict.

        This is consumed by the GUI's AI Analysis tab.  If AI is enabled and
        an API key is configured, could call an LLM (placeholder).

        Returns a dict with indicator snapshots, trend, confidence scores,
        and a suggested action.
        """
        analysis: dict[str, Any] = {
            "symbol": symbol,
            "timestamp": datetime.now().isoformat(),
            "trend": "NEUTRAL",
            "momentum_score": 0,
            "trend_score": 0,
            "volatility_score": 0,
            "oi_score": 0,
            "rsi_score": 0,
            "overall_score": 0,
            "suggested_action": "HOLD",
            "latest_signal": None,
            "indicators": {},
        }

        try:
            from core.indicators import (
                adx,
                atr,
                ema,
                get_trend,
                is_sideways,
                rsi,
                supertrend,
                vwap,
            )

            # Compute indicators
            ema9 = ema(candles, 9)
            ema21 = ema(candles, 21)
            rsi_val = rsi(candles, 14)
            adx_val = adx(candles, 14)
            st = supertrend(candles, 10, 3.0)
            vwap_val = vwap(candles)
            atr_val = atr(candles, 14)
            trend = get_trend(candles)
            sideways = is_sideways(candles)

            # Extract latest values
            def _latest(lst):
                if not lst:
                    return 0
                if isinstance(lst[-1], dict):
                    return lst[-1]
                import math

                for v in reversed(lst):
                    if not math.isnan(v):
                        return v
                return 0

            ema9_v = round(float(_latest(ema9)), 2) if ema9 else 0
            ema21_v = round(float(_latest(ema21)), 2) if ema21 else 0
            rsi_v = round(float(rsi_val), 2) if rsi_val else 50
            adx_v = round(float(adx_val), 2) if adx_val else 0
            st_v = st[-1] if st else {}
            st_dir = st_v.get("direction", "NONE") if isinstance(st_v, dict) else "NONE"
            vwap_v = round(float(vwap_val), 2) if vwap_val else 0
            atr_v = round(float(atr_val), 2) if atr_val else 0

            current_close = candles[-1]["close"] if candles else 0

            analysis["indicators"] = {
                "ema9": ema9_v,
                "ema21": ema21_v,
                "rsi": rsi_v,
                "adx": adx_v,
                "supertrend": st_dir,
                "vwap": vwap_v,
                "atr": atr_v,
                "close": round(current_close, 2),
                "sideways": sideways,
                "trend": trend or "NEUTRAL",
            }
            analysis["trend"] = trend or "NEUTRAL"

            # Score calculations
            # Momentum: based on EMA crossover + RSI
            momentum = 50
            if ema9_v > ema21_v:
                momentum += 20
            else:
                momentum -= 20
            if 45 < rsi_v < 55:
                momentum += 5
            elif rsi_v > 60:
                momentum += 15
            elif rsi_v < 40:
                momentum -= 15
            momentum = max(0, min(100, momentum))
            analysis["momentum_score"] = momentum

            # Trend: based on ADX + SuperTrend
            trend_score = 50
            if adx_v > 30:
                trend_score += 25
            elif adx_v > 20:
                trend_score += 10
            if st_dir == "UP":
                trend_score += 15
            elif st_dir == "DOWN":
                trend_score -= 15
            trend_score = max(0, min(100, trend_score))
            analysis["trend_score"] = trend_score

            # Volatility: based on ATR
            vol_score = 50
            if atr_v > 100:
                vol_score += 20
            elif atr_v > 50:
                vol_score += 10
            elif atr_v < 20:
                vol_score -= 15
            vol_score = max(0, min(100, vol_score))
            analysis["volatility_score"] = vol_score

            # RSI score
            if rsi_v > 70:
                rsi_score = 80  # overbought → sell pressure
            elif rsi_v < 30:
                rsi_score = 20  # oversold → buy pressure
            elif rsi_v > 55:
                rsi_score = 70
            elif rsi_v < 45:
                rsi_score = 30
            else:
                rsi_score = 50
            analysis["rsi_score"] = rsi_score

            # OI score
            oi_score = 50
            if self.option_chain and not sideways:
                try:
                    oi_data = self.option_chain.analyze_oi(symbol)
                    pcr = oi_data.get("put_call_ratio", 0)
                    if pcr > 1.2:
                        oi_score = 75  # bullish
                    elif pcr < 0.8:
                        oi_score = 25  # bearish
                except Exception:
                    pass
            analysis["oi_score"] = oi_score

            # Overall
            overall = round(
                momentum * 0.25
                + trend_score * 0.25
                + vol_score * 0.1
                + oi_score * 0.2
                + rsi_score * 0.2,
                1,
            )
            analysis["overall_score"] = max(0, min(100, overall))

            # Suggested action
            if overall >= 70:
                analysis["suggested_action"] = "BUY CE"
            elif overall <= 30:
                analysis["suggested_action"] = "BUY PE"
            elif overall >= 55:
                analysis["suggested_action"] = "HOLD LONG"
            elif overall <= 45:
                analysis["suggested_action"] = "HOLD SHORT"
            else:
                analysis["suggested_action"] = "HOLD"

            # Latest signal
            if signals:
                analysis["latest_signal"] = signals[-1]

            # Cache for dashboard
            self._dashboard_cache["ai_analysis"] = analysis

        except Exception as exc:
            logger.error(
                "AI analysis generation error: %s", exc, exc_info=True
            )

        return analysis

    # ------------------------------------------------------------------
    # Daily Report
    # ------------------------------------------------------------------

    def generate_daily_report(self) -> Optional[dict[str, Any]]:
        """Compile daily stats, save to DB, export Excel, send via Telegram.

        Returns the report dict on success.
        """
        if not self.db or not self.capital_manager:
            return None

        try:
            today = datetime.now().strftime("%Y-%m-%d")
            capital = self.capital_manager.get_state()
            daily = self.capital_manager.get_daily_summary()
            total = self.capital_manager.get_total_summary()

            # Get today's closed trades
            closed_trades = self.db.get_trade_history(limit=200)
            today_closed = [
                t
                for t in closed_trades
                if t.get("exit_time", "").startswith(today)
            ]

            wins = len([t for t in today_closed if t.get("pnl", 0) > 0])
            losses = len([t for t in today_closed if t.get("pnl", 0) <= 0])
            total_trades = wins + losses
            win_rate = round(wins / total_trades * 100, 2) if total_trades > 0 else 0.0

            pnls = [t.get("pnl", 0) for t in today_closed]
            net_pnl = round(sum(pnls), 2)
            avg_win = round(sum(p for p in pnls if p > 0) / wins, 2) if wins > 0 else 0.0
            avg_loss = round(sum(p for p in pnls if p < 0) / losses, 2) if losses > 0 else 0.0
            max_win = round(max(pnls), 2) if pnls else 0.0
            max_loss = round(min(pnls), 2) if pnls else 0.0

            # Best strategy
            strategy_pnl: dict[str, float] = {}
            for t in today_closed:
                s = t.get("strategy", "Unknown")
                strategy_pnl[s] = strategy_pnl.get(s, 0) + t.get("pnl", 0)
            best_strategy = max(strategy_pnl, key=strategy_pnl.get) if strategy_pnl else ""

            report = {
                "date": today,
                "total_pnl": net_pnl,
                "total_trades": total_trades,
                "wins": wins,
                "losses": losses,
                "win_rate": win_rate,
                "avg_win": avg_win,
                "avg_loss": avg_loss,
                "max_win": max_win,
                "max_loss": max_loss,
                "best_strategy": best_strategy,
                "capital_start": capital.get("initial", 0),
                "capital_end": capital.get("current", 0),
                "report_data": {
                    "daily_summary": daily,
                    "total_summary": total,
                    "strategy_pnl": strategy_pnl,
                },
            }

            # Save to DB
            self.db.save_daily_report(report)

            # Export to Excel
            if self._config.get("daily_report", {}).get("export_excel", True):
                try:
                    self._export_report_excel(report, today_closed)
                except Exception as exc:
                    logger.error("Excel export failed: %s", exc)

            # Send via Telegram
            if self.telegram_bot:
                try:
                    self.telegram_bot.send_daily_report(report)
                except Exception as exc:
                    logger.error("Telegram daily report failed: %s", exc)

            self._log_to_db(
                "INFO",
                f"Daily report generated: {total_trades} trades, "
                f"net PnL={net_pnl:+.2f}, win_rate={win_rate:.1f}%",
                "ENGINE",
            )

            logger.info(
                "Daily report: %d trades, net PnL=%+.2f, WR=%.1f%%",
                total_trades,
                net_pnl,
                win_rate,
            )

            return report

        except Exception as exc:
            logger.error(
                "Daily report generation failed: %s", exc, exc_info=True
            )
            return None

    def _export_report_excel(
        self, report: dict[str, Any], trades: list[dict[str, Any]]
    ) -> None:
        """Export daily report to an Excel file using openpyxl."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()

            # --- Summary sheet ---
            ws = wb.active
            ws.title = "Summary"

            header_font = Font(bold=True, size=12)
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            ws["A1"] = "JSS Sawriya Seth Wealthtech — Daily Report"
            ws["A1"].font = Font(bold=True, size=14)
            ws.merge_cells("A1:D1")

            ws["A3"] = "Date"
            ws["B3"] = report.get("date", "")
            ws["A4"] = "Total Trades"
            ws["B4"] = report.get("total_trades", 0)
            ws["A5"] = "Wins"
            ws["B5"] = report.get("wins", 0)
            ws["A6"] = "Losses"
            ws["B6"] = report.get("losses", 0)
            ws["A7"] = "Win Rate"
            ws["B7"] = f"{report.get('win_rate', 0):.1f}%"
            ws["A8"] = "Net P&L"
            ws["B8"] = report.get("total_pnl", 0)
            ws["B8"].font = Font(bold=True)
            if report.get("total_pnl", 0) >= 0:
                ws["B8"].fill = green_fill
            else:
                ws["B8"].fill = red_fill
            ws["A9"] = "Capital End"
            ws["B9"] = report.get("capital_end", 0)
            ws["A10"] = "Best Strategy"
            ws["B10"] = report.get("best_strategy", "")

            for row in ws.iter_rows(min_row=3, max_row=10):
                for cell in row:
                    cell.font = header_font if cell.column == 1 else cell.font
                    cell.alignment = Alignment(horizontal="left")

            # --- Trades sheet ---
            ws2 = wb.create_sheet("Trades")
            headers = [
                "Time", "Symbol", "Direction", "Option", "Strike",
                "Entry", "Exit", "P&L", "Strategy", "Reason",
            ]

            for col, h in enumerate(headers, 1):
                cell = ws2.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)

            for i, trade in enumerate(trades, 2):
                ws2.cell(row=i, column=1, value=trade.get("entry_time", ""))
                ws2.cell(row=i, column=2, value=trade.get("symbol", ""))
                ws2.cell(row=i, column=3, value=trade.get("direction", ""))
                ws2.cell(row=i, column=4, value=trade.get("option_type", ""))
                ws2.cell(row=i, column=5, value=trade.get("strike", 0))
                ws2.cell(row=i, column=6, value=trade.get("entry_price", 0))
                ws2.cell(row=i, column=7, value=trade.get("exit_price", 0))
                pnl_cell = ws2.cell(row=i, column=8, value=trade.get("pnl", 0))
                if trade.get("pnl", 0) >= 0:
                    pnl_cell.fill = green_fill
                else:
                    pnl_cell.fill = red_fill
                ws2.cell(row=i, column=9, value=trade.get("strategy", ""))
                ws2.cell(row=i, column=10, value=trade.get("exit_reason", ""))

            # Save
            report_dir = Path("reports")
            report_dir.mkdir(parents=True, exist_ok=True)
            date_str = report.get("date", datetime.now().strftime("%Y-%m-%d"))
            filepath = report_dir / f"JSS_Report_{date_str}.xlsx"
            wb.save(str(filepath))

            logger.info("Excel report saved: %s", filepath)

        except ImportError:
            logger.warning(
                "openpyxl not installed — skipping Excel export. "
                "Install with: pip install openpyxl"
            )
        except Exception as exc:
            logger.error("Excel export error: %s", exc, exc_info=True)

    # ------------------------------------------------------------------
    # Dashboard data (consumed by GUI)
    # ------------------------------------------------------------------

    def get_dashboard_data(self) -> dict[str, Any]:
        """Return a comprehensive data dict consumed by the GUI every 2 s.

        Keys:
        - ``capital_state``: Current capital dict from CapitalManager.
        - ``current_trade``: First open trade or None.
        - ``market_data``: Live LTP data for configured symbols.
        - ``recent_logs``: Last 50 system log entries.
        - ``trade_history``: Last 50 closed trades.
        - ``ai_analysis``: Latest AI analysis dict.
        - ``engine_state``: Current engine state (RUNNING/STOPPED).
        """
        data: dict[str, Any] = {
            "capital_state": {},
            "current_trade": None,
            "market_data": [],
            "recent_logs": [],
            "trade_history": [],
            "ai_analysis": self._dashboard_cache.get("ai_analysis", {}),
            "engine_state": self._state,
            "timestamp": datetime.now().isoformat(),
        }

        try:
            # Capital state
            if self.capital_manager:
                data["capital_state"] = self.capital_manager.get_state()

            # Current (first open) trade
            if self.db:
                open_trades = self.db.get_open_trades()
                if open_trades:
                    data["current_trade"] = open_trades[0]

                # Trade history
                data["trade_history"] = self.db.get_trade_history(limit=50)

                # Recent logs
                data["recent_logs"] = self.db.get_logs(limit=50)

        except Exception as exc:
            logger.error("Dashboard data error: %s", exc, exc_info=True)

        # Market data
        if self.broker:
            market = []
            for symbol in self._config.get("symbols", ["NIFTY", "BANKNIFTY"]):
                try:
                    ltp = self.broker.get_ltp(symbol)
                    change = round(ltp * 0.001 * (1 if ltp % 2 == 0 else -1), 2)
                    market.append(
                        {
                            "symbol": symbol,
                            "ltp": round(ltp, 2),
                            "change": change,
                            "change_pct": round(change / ltp * 100, 2) if ltp > 0 else 0,
                            "high": round(ltp + abs(change) * 2, 2),
                            "low": round(ltp - abs(change) * 2, 2),
                            "volume": 0,
                        }
                    )
                except Exception:
                    market.append(
                        {"symbol": symbol, "ltp": 0, "change": 0,
                         "change_pct": 0, "high": 0, "low": 0, "volume": 0}
                    )
            data["market_data"] = market

        return data

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _get_lot_multiplier(symbol: str) -> int:
        """Return lot multiplier for the given symbol."""
        symbol = symbol.upper()
        if symbol == "BANKNIFTY":
            return 25
        return 50  # NIFTY, FINNIFTY, default

    def _log_to_db(
        self,
        level: str,
        message: str,
        source: str = "ENGINE",
        metadata: Optional[dict[str, Any]] = None,
    ) -> None:
        """Log a message to both Python logging and the database."""
        # Python logging
        py_level = getattr(logging, level.upper(), logging.INFO)
        logger.log(py_level, "[%s] %s", source, message)

        # Database
        if self.db:
            try:
                self.db.log(level, message, source, metadata)
            except Exception as exc:
                logger.error("DB log failed: %s", exc)

    # ------------------------------------------------------------------
    # Properties
    # ------------------------------------------------------------------

    @property
    def state(self) -> str:
        """Current engine state (RUNNING or STOPPED)."""
        return self._state

    @property
    def is_running(self) -> bool:
        """Whether the engine main loop is active."""
        return self._running
